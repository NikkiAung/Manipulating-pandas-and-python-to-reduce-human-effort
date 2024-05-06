import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import os

# Initialize the array to store the names of the generated text files
txt_files = []

def browse_excel_file():
    global excel_file_path
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, excel_file_path)

def browse_output_path():
    global csv_output_path
    csv_output_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    output_path_entry.delete(0, tk.END)
    output_path_entry.insert(0, csv_output_path)

def convert_to_csv():
    try:
        df = pd.read_excel(excel_file_path)
        df.to_csv(csv_output_path, index=False)
        status_label.config(text="Excel file converted to CSV and saved successfully", fg="green")
        
        # Generate RBQL template file
        rbql_template_path = os.path.splitext(csv_output_path)[0] + "_rbql.txt"
        with open(rbql_template_path, 'w') as f:
            f.write("# Write your RBQL query here")
        status_label.config(text=f"RBQL template file generated: {rbql_template_path}", fg="green")
        
        # Fill the template with data from the CSV file
        fill_template(rbql_template_path, df)
        
        # Define paths to Excel template and output Excel file
        excel_template_path = '/Users/aungnandaoo/Desktop/ExcelData/Template/PLA label template.xlsx'
        output_dir = '/Users/aungnandaoo/Desktop/ExcelData/NewTemplate'
        
        # Fill Excel template with data from text files
        for text_file_path in txt_files:
            data = read_text_data(text_file_path)
            fill_excel_template(data, excel_template_path, output_dir)
            
    except Exception as e:
        status_label.config(text=f"Error: {str(e)}", fg="red")

def fill_template(template_path, df):
    global txt_files
    for index, row in df.iterrows():
        batch_lot_number = row['Batch Lot Number']
        template_file_path = f"{os.path.splitext(template_path)[0]}_{batch_lot_number}_template.txt"
        txt_files.append(template_file_path)
        with open(template_file_path, 'w') as template_file:
            # textPart
            template_file.write(f"batch_lot_number = '{batch_lot_number}'\n")
            template_file.write(f"vendor = '{row.get('Vendor', '-')}'\n")
            template_file.write(f"color = '{row.get('Color', '-')}'\n")
            template_file.write(f"color_formulation_number = '{row.get('Color Formulation Number', '-')}'\n")
            template_file.write(f"illuminant = '{row.get('Illuminant', '-')}'\n")
            template_file.write(f"submit_type = '{row.get('Submit Type', '-')}'\n")
            template_file.write(f"material_name_and_grade = '{row.get('Material Name and Grade', '-')}'\n")
            template_file.write(f"program = '{row.get('Program', '-')}'\n")
            template_file.write(f"vendor_visual_assessment = '{row.get('Vendor Visual Assessment', '-')}'\n")
            template_file.write(f"resin_production_date = '{row.get('Resin Production Date', '-')}'\n")
            template_file.write(f"resin_delivery_date = '{row.get('Resin Delivery Date', '-')}'\n")
            template_file.write(f"color_measurement_date = '{row.get('Color Measurement Date', '-')}'\n")
            template_file.write(f"measurement_location = '{row.get('Measurement location', '-')}'\n")
            template_file.write(f"approval_status = '{row.get('Approval Status', '-')}'\n")
            # Checking and replacing NaN values with '-'
            id_review_date = row.get('ID Review Date', '-')
            id_review_date = '-' if pd.isna(id_review_date) else id_review_date
            template_file.write(f"id_review_date = '{id_review_date}'\n")
            # numberPart
            template_file.write(f"l_star = '{round(float(row['L*']), 2) if pd.notnull(row['L*']) and str(row['L*']) != '-' else '-'}'\n")
            template_file.write(f"a_star = '{round(float(row['a*']), 2) if pd.notnull(row['a*']) and str(row['a*']) != '-' else '-'}'\n")
            template_file.write(f"b_star = '{round(float(row['b*']), 2) if pd.notnull(row['b*']) and str(row['b*']) != '-' else '-'}'\n")
            template_file.write(f"c_star_f2 = '{round(float(row['C*  F2']), 2) if pd.notnull(row['C*  F2']) and str(row['C*  F2']) != '-' else '-'}'\n")
            template_file.write(f"h_f2 = '{round(float(row['h F2']), 2) if pd.notnull(row['h F2']) and str(row['h F2']) != '-' else '-'}'\n")
            template_file.write(f"d_l_star = '{round(float(row['dL*']), 2) if pd.notnull(row['dL*']) and str(row['dL*']) != '-' else '-'}'\n")
            template_file.write(f"d_a_star = '{round(float(row['da*']), 2) if pd.notnull(row['da*']) and str(row['da*']) != '-' else '-'}'\n")
            template_file.write(f"d_b_star = '{round(float(row['db*']), 2) if pd.notnull(row['db*']) and str(row['db*']) != '-' else '-'}'\n")
            template_file.write(f"de94 = '{round(float(row['DE94']), 2) if pd.notnull(row['DE94']) and str(row['DE94']) != '-' else '-'}'\n")
            template_file.write(f"d_c_star = '{round(float(row['dC*']), 2) if pd.notnull(row['dC*']) and str(row['dC*']) != '-' else '-'}'\n")
            template_file.write(f"d_h = '{round(float(row['dH']), 2) if pd.notnull(row['dH']) and str(row['dH']) != '-' else '-'}'\n")
            template_file.write(f"c_star_std = '{round(float(row['c* std']), 2) if pd.notnull(row['c* std']) and str(row['c* std']) != '-' else '-'}'\n")
            template_file.write(f"h_std = '{round(float(row['h std']), 2) if pd.notnull(row['h std']) and str(row['h std']) != '-' else '-'}'\n")
            template_file.write(f"additional_notes = '{row.get('Additional Notes', '-')}'\n")
            additional_notes = row.get('Additional Notes', '-')
            additional_notes = '-' if pd.isna(additional_notes) else additional_notes
            template_file.write(f"additional_notes = '{additional_notes}'\n")
            
        status_label.config(text=f"Template filled for Batch Lot Number: {batch_lot_number}", fg="green")

def read_text_data(file_path):
    with open(file_path, 'r') as file:
        data = {}
        for line in file:
            key, value = line.strip().split(' = ')
            data[key.strip()] = value.strip().strip("'")
    return data

def fill_excel_template(data, template_path, output_dir):
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Define the cell positions for filling data
    cell_positions = {
        "D4": data["vendor"],
        "E4": data["color"],
        "F4": data["color_formulation_number"],
        "D6": data["illuminant"],
        "E6": data["submit_type"],
        "F6": data["material_name_and_grade"],
        "D8": data["program"],
        "E8": data["batch_lot_number"],
        "F8": data["vendor_visual_assessment"],
        "D10": data["resin_production_date"],
        "E10": data["color_measurement_date"],
        "F10": data["resin_delivery_date"],
        "D12": data["measurement_location"],
        "E12": data["approval_status"],
        "F12": data["id_review_date"],
        "D14": data["l_star"],
        "E14": data["a_star"],
        "F14": data["b_star"],
        "D16": data["h_f2"],
        "E16": data["h_f2"],
        "F16": data["d_l_star"],
        "D18": data["d_a_star"],
        "E18": data["d_b_star"],
        "F18": data["de94"],
        "D20": data["d_c_star"],
        "E20": data["d_h"],
        "D22": data["c_star_std"],
        "E22": data["h_std"],
        "F22": data["additional_notes"]
    }
    
    # Create the output file name based on the batch lot number
    output_filename = f"outputfile_{data['batch_lot_number']}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    # Fill in the data
    for cell, value in cell_positions.items():
        ws[cell] = value
    
    wb.save(output_path)

# Create the main window
root = tk.Tk()
root.title("Excel to CSV Converter")

# Create and place widgets
excel_file_label = tk.Label(root, text="Excel File:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(root, width=50)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

browse_excel_button = tk.Button(root, text="Browse", command=browse_excel_file)
browse_excel_button.grid(row=0, column=2, padx=5, pady=5)

output_path_label = tk.Label(root, text="CSV Output Path:")
output_path_label.grid(row=1, column=0, padx=5, pady=5)

output_path_entry = tk.Entry(root, width=50)
output_path_entry.grid(row=1, column=1, padx=5, pady=5)

browse_output_button = tk.Button(root, text="Browse", command=browse_output_path)
browse_output_button.grid(row=1, column=2, padx=5, pady=5)

convert_button = tk.Button(root, text="Convert to CSV", command=convert_to_csv)
convert_button.grid(row=2, column=1, padx=5, pady=5)

status_label = tk.Label(root, text="", fg="green")
status_label.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

# Start the GUI event loop
root.mainloop()

# Print the generated text files
print("Generated Text Files:")
for text_file in txt_files:
    print(text_file)
txt_files.clear()

#Error 'float' object has no attribute 'fillna'